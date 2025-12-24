"""Excel report export service using openpyxl."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..schemas.results import AnalysisResponse, ReportEntry


def export_analysis_to_excel(
    analysis: AnalysisResponse,
    selected_columns: List[str] | None = None,
    timestamp_column: str | None = None,
) -> bytes:
    """
    Export analysis results to Excel format.
    
    Args:
        analysis: Analysis response containing report data
        selected_columns: List of selected column names
        timestamp_column: Name of timestamp column
        
    Returns:
        Excel file content as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    _create_overview_sheet(wb, analysis, selected_columns, timestamp_column)
    _create_indicators_sheet(wb, analysis, selected_columns)
    _create_forecast_sheet(wb, analysis, selected_columns)
    _create_trend_chart_sheet(wb, analysis, selected_columns)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes


def _create_overview_sheet(
    wb: Workbook,
    analysis: AnalysisResponse,
    selected_columns: List[str] | None,
    timestamp_column: str | None,
) -> None:
    """Create overview sheet with metadata."""
    ws = wb.create_sheet("数据概览", 0)
    
    # Title
    ws['A1'] = "拜耳法氧化铝生产分析报告"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:B1')
    
    # Generation time
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A2'] = "生成时间"
    ws['B2'] = generated_at
    
    # Metadata
    metadata = analysis.metadata or {}
    row = 4
    
    ws[f'A{row}'] = "数据统计"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="2C3E50")
    row += 1
    
    if metadata.get("row_count"):
        ws[f'A{row}'] = "记录数"
        ws[f'B{row}'] = metadata["row_count"]
        row += 1
    
    if metadata.get("column_count"):
        ws[f'A{row}'] = "字段数"
        ws[f'B{row}'] = metadata["column_count"]
        row += 1
    
    if metadata.get("missing_ratio") is not None:
        ws[f'A{row}'] = "缺失率"
        ws[f'B{row}'] = f"{metadata['missing_ratio'] * 100:.2f}%"
        row += 1
    
    effective_timestamp = (
        timestamp_column
        or analysis.agent_payload.timestamp_column
        or metadata.get("timestamp_column")
        or "自动识别"
    )
    ws[f'A{row}'] = "时间字段"
    ws[f'B{row}'] = effective_timestamp
    row += 2
    
    # Selected columns
    if selected_columns:
        ws[f'A{row}'] = "分析字段"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="2C3E50")
        row += 1
        for col in selected_columns:
            ws[f'A{row}'] = col
            row += 1
    
    # Style column A
    for cell in ws['A']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40


def _create_indicators_sheet(
    wb: Workbook,
    analysis: AnalysisResponse,
    selected_columns: List[str] | None,
) -> None:
    """Create indicators detail sheet."""
    ws = wb.create_sheet("指标详情")
    
    # Headers
    headers = [
        "参数名称",
        "当前状态",
        "趋势",
        "预测趋势",
        "最新值",
        "均值",
        "标准差",
        "预测解读",
        "操作建议",
        "主要相关因子",
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    effective_columns = selected_columns or []
    row_idx = 2
    
    for entry in analysis.report:
        if effective_columns and entry.parameter not in effective_columns:
            continue
        
        ws.cell(row=row_idx, column=1, value=entry.parameter)
        ws.cell(row=row_idx, column=2, value=entry.current_status)
        ws.cell(row=row_idx, column=3, value=entry.trend)
        ws.cell(row=row_idx, column=4, value=entry.forecast_trend)
        ws.cell(row=row_idx, column=5, value=_format_value(entry.latest_value))
        ws.cell(row=row_idx, column=6, value=_format_value(entry.mean_value))
        ws.cell(row=row_idx, column=7, value=_format_value(entry.std_dev))
        ws.cell(row=row_idx, column=8, value=entry.forecast_summary or "")
        ws.cell(row=row_idx, column=9, value="; ".join(entry.recommendations) if entry.recommendations else "")
        ws.cell(row=row_idx, column=10, value=", ".join(entry.influence_factors) if entry.influence_factors else "")
        
        row_idx += 1
    
    # Apply borders and alignment
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'),
    )
    
    for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Adjust column widths
    column_widths = [15, 10, 10, 10, 12, 12, 12, 30, 35, 25]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _create_forecast_sheet(
    wb: Workbook,
    analysis: AnalysisResponse,
    selected_columns: List[str] | None,
) -> None:
    """Create forecast summary sheet."""
    ws = wb.create_sheet("趋势摘要")
    
    # Headers
    ws['A1'] = "参数名称"
    ws['B1'] = "趋势摘要"
    
    for cell in [ws['A1'], ws['B1']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    effective_columns = selected_columns or []
    row_idx = 2
    
    for name, summary in analysis.forecast.summary.items():
        if effective_columns and name not in effective_columns:
            continue
        
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=summary)
        row_idx += 1
    
    # Apply borders
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'),
    )
    
    for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _create_trend_chart_sheet(
    wb: Workbook,
    analysis: AnalysisResponse,
    selected_columns: List[str] | None,
) -> None:
    """Create sheet with historical trend charts."""
    ws = wb.create_sheet("历史趋势")

    history_tail = analysis.forecast.history_tail or {}
    if not history_tail:
        ws['A1'] = "暂无历史数据"
        return

    effective_columns = selected_columns or list(history_tail.keys())
    seen_any = False
    row_offset = 1

    for parameter in effective_columns:
        points = history_tail.get(parameter)
        if not points:
            continue

        seen_any = True

        # Write data table
        ws.cell(row=row_offset, column=1, value=f"参数：{parameter}")
        ws.cell(row=row_offset + 1, column=1, value="时间")
        ws.cell(row=row_offset + 1, column=2, value="数值")

        data_start_row = row_offset + 2
        data_end_row = data_start_row + len(points) - 1

        for idx, point in enumerate(points, start=data_start_row):
            ws.cell(row=idx, column=1, value=point.index)
            ws.cell(row=idx, column=2, value=point.value)

        # Create chart
        chart = LineChart()
        chart.title = f"{parameter} 历史趋势"
        chart.style = 13
        chart.y_axis.title = parameter
        chart.x_axis.title = "时间"

        data_ref = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 8
        chart.width = 20

        ws.add_chart(chart, f"E{row_offset + 1}")

        row_offset = data_end_row + 3

    if not seen_any:
        ws['A1'] = "选定指标缺少历史数据"


def _format_value(value: float | None) -> str:
    """Format numeric value for display."""
    if value is None:
        return "—"
    return f"{value:.4f}"
